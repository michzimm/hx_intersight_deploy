#!/usr/bin/env python
import csv
import sys
import json
import os.path
import argparse
import getpass
import intersight
import logging
import datetime
import signal
from time import sleep
from openpyxl import load_workbook
from colorama import Fore, Back, Style
from beautifultable import BeautifulTable
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from intersight.intersight_api_client import IntersightApiClient
from intersight.apis import asset_device_claim_api
from intersight.rest import ApiException
from intersight.apis import hyperflex_local_credential_policy_api
from intersight.apis import hyperflex_sys_config_policy_api
from intersight.apis import hyperflex_vcenter_config_policy_api
from intersight.apis import hyperflex_cluster_storage_policy_api
from intersight.apis import hyperflex_auto_support_policy_api
from intersight.apis import hyperflex_node_config_policy_api
from intersight.models import hyperflex_ip_addr_range
from intersight.apis import hyperflex_cluster_network_policy_api
from intersight.models import hyperflex_named_vlan
from intersight.apis import hyperflex_proxy_setting_policy_api
from intersight.apis import hyperflex_cluster_profile_api
from intersight.apis import compute_rack_unit_api
from intersight.apis import hyperflex_node_profile_api
from intersight.apis import hyperflex_software_version_policy_api
from intersight.models import hyperflex_mac_addr_prefix_range
import source.device_connector
from imcsdk.imchandle import ImcHandle

class InputRecord(object):
    def __init__(self, hx_profile_name=None, cluster_type=None, hxdp_version=None, description=None, data_vlan_id=None, cluster_mgmt_ip=None, mac_address_prefix=None, hx_nodes_cimc_ips=None, hx_nodes_cimc_user=None, local_credential_policy_name=None, hypervisor_admin_user=None, sys_config_policy_name=None, dns_suffix=None, timezone=None, dns_servers=None, ntp_servers=None, vcenter_policy_name=None, vcenter=None, vcenter_user=None, vcenter_dc=None, vcenter_sso=None, cluster_storage_policy_name=None, vdi_optimization=None, clean_partitions=None, auto_support_policy_name=None, auto_support=None, auto_support_email=None, node_config_policy_name=None, hostname_prefix=None, mgmt_start_ip=None, mgmt_end_ip=None, mgmt_subnet_mask=None, mgmt_gw=None, cont_vm_start_ip=None, cont_vm_end_ip=None, cont_vm_subnet_mask=None, cont_vm_gw=None, cluster_network_policy_name=None, mgmt_vlan_id=None, uplink_speed=None, jumbo_frames=None, proxy_setting_policy_name=None, proxy_hostname=None, proxy_port=None, proxy_username=None, proxy_password=None):

        if hx_profile_name is not None:
            self.hx_profile_name = hx_profile_name
        if cluster_type is not None:
            self.cluster_type = cluster_type
        if hxdp_version is not None:
            self.hxdp_version = hxdp_version
        if description is not None:
            self.description = description
        if data_vlan_id is not None:
            self.data_vlan_id = data_vlan_id
        if cluster_mgmt_ip is not None:
            self.cluster_mgmt_ip = cluster_mgmt_ip
        if mac_address_prefix is not None:
            self.mac_address_prefix = mac_address_prefix
        if hx_nodes_cimc_ips is not None:
            self.hx_nodes_cimc_ips = hx_nodes_cimc_ips
        if hx_nodes_cimc_user is not None:
            self.hx_nodes_cimc_user = hx_nodes_cimc_user
        if local_credential_policy_name is not None:
            self.local_credential_policy_name = local_credential_policy_name
        if hypervisor_admin_user is not None:
            self.hypervisor_admin_user = hypervisor_admin_user
        if sys_config_policy_name is not None:
            self.sys_config_policy_name = sys_config_policy_name
        if dns_suffix is not None:
            self.dns_suffix = dns_suffix
        if timezone is not None:
            self.timezone = timezone
        if dns_servers is not None:
            self.dns_servers = dns_servers
        if ntp_servers is not None:
            self.ntp_servers = ntp_servers
        if vcenter_policy_name is not None:
            self.vcenter_policy_name = vcenter_policy_name
        if vcenter is not None:
            self.vcenter = vcenter
        if vcenter_user is not None:
            self.vcenter_user = vcenter_user
        if vcenter_dc is not None:
            self.vcenter_dc = vcenter_dc
        if vcenter_sso is not None:
            self.vcenter_sso = vcenter_sso
        if cluster_storage_policy_name is not None:
            self.cluster_storage_policy_name = cluster_storage_policy_name
        if vdi_optimization is not None:
            self.vdi_optimization = vdi_optimization
        if clean_partitions is not None:
            self.clean_partitions = clean_partitions
        if auto_support_policy_name is not None:
            self.auto_support_policy_name = auto_support_policy_name
        if auto_support is not None:
            self.auto_support = auto_support
        if auto_support_email is not None:
            self.auto_support_email = auto_support_email
        if node_config_policy_name is not None:
            self.node_config_policy_name = node_config_policy_name
        if hostname_prefix is not None:
            self.hostname_prefix = hostname_prefix
        if mgmt_start_ip is not None:
            self.mgmt_start_ip = mgmt_start_ip
        if mgmt_end_ip is not None:
            self.mgmt_end_ip = mgmt_end_ip
        if mgmt_subnet_mask is not None:
            self.mgmt_subnet_mask = mgmt_subnet_mask
        if mgmt_gw is not None:
            self.mgmt_gw = mgmt_gw
        if cont_vm_start_ip is not None:
            self.cont_vm_start_ip = cont_vm_start_ip
        if cont_vm_end_ip is not None:
            self.cont_vm_end_ip = cont_vm_end_ip
        if cont_vm_subnet_mask is not None:
            self.cont_vm_subnet_mask = cont_vm_subnet_mask
        if cont_vm_gw is not None:
            self.cont_vm_gw = cont_vm_gw
        if cluster_network_policy_name is not None:
            self.cluster_network_policy_name = cluster_network_policy_name
        if mgmt_vlan_id is not None:
            self.mgmt_vlan_id = mgmt_vlan_id
        if uplink_speed is not None:
            self.uplink_speed = uplink_speed
        if jumbo_frames is not None:
            self.jumbo_frames = jumbo_frames
        if proxy_setting_policy_name is not None:
            self.proxy_setting_policy_name = proxy_setting_policy_name
        if proxy_hostname is not None:
            self.proxy_hostname = proxy_hostname
        if proxy_port is not None:
            self.proxy_port = proxy_port
        if proxy_username is not None:
            self.proxy_username = proxy_username
        if proxy_password is not None:
            self.proxy_password = proxy_password


    def __str__(self):
        return str(self.__class__) + ": " + str(self.__dict__)

####################################################################################################
#
#
# FUNCTIONS
#
#
####################################################################################################


def create_software_version_policy(api_instance, policy_name, hxdp_version):

    # Get API handle
    software_version_policy_handle = hyperflex_software_version_policy_api.HyperflexSoftwareVersionPolicyApi(api_instance)

    # Setup variables

    # Create API body
    software_version_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'HxdpVersion':hxdp_version
    }

    # Execute API
    try:
        api_response = software_version_policy_handle.hyperflex_software_version_policies_post(software_version_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_credential_policy(api_instance, policy_name, hypervisor_admin_user, hypervisor_password, controller_password):

    # Get API handle
    credential_policy_handle = hyperflex_local_credential_policy_api.HyperflexLocalCredentialPolicyApi(api_instance)

    # Setup variables

    # Create API body
    credential_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'FactoryHypervisorPassword': True,
        'HypervisorAdmin':hypervisor_admin_user,
        'HypervisorAdminPwd':hypervisor_password,
        'HxdpRootPwd':controller_password
    }

    # Execute API
    try:
        api_response = credential_policy_handle.hyperflex_local_credential_policies_post(credential_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_sys_config_policy(api_instance, policy_name, timezone, dns_suffix, dns_servers, ntp_servers):

    # Get API handle
    sys_config_policy_handle = hyperflex_sys_config_policy_api.HyperflexSysConfigPolicyApi(api_instance)

    # Setup variables
    dns_servers_list = []
    dns_servers_array = (dns_servers).split(',')
    for dns_server in dns_servers_array:
        dns_servers_list.append(dns_server)
    ntp_servers_list = []
    ntp_servers_array = (ntp_servers).split(',')
    for ntp_server in ntp_servers_array:
        ntp_servers_list.append(ntp_server)

    # Create API body
    sys_config_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'Timezone':timezone,
        'DnsDomainName':dns_suffix,
        'DnsServers':dns_servers_list,
        'NtpServers':ntp_servers_list
    }

    # Execute API
    try:
        api_response = sys_config_policy_handle.hyperflex_sys_config_policies_post(sys_config_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_vcenter_policy(api_instance, policy_name, vcenter, vcenter_user, vcenter_password, vcenter_dc, vcenter_sso=''):

    # Get API handle
    vcenter_config_policy_handle = hyperflex_vcenter_config_policy_api.HyperflexVcenterConfigPolicyApi(api_instance)

    # Setup variables

    # Create API body
    vcenter_config_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'Hostname':vcenter,
        'Username':vcenter_user,
        'Password':vcenter_password,
        'DataCenter':vcenter_dc,
        'SsoUrl': vcenter_sso
    }

    # Execute API
    try:
        api_response = vcenter_config_policy_handle.hyperflex_vcenter_config_policies_post(vcenter_config_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_cluster_storage_policy(api_instance, policy_name, vdi_optimization, clean_partitions):

    # Get API handle
    cluster_storage_policy_handle = hyperflex_cluster_storage_policy_api.HyperflexClusterStoragePolicyApi(api_instance)

    # Setup variables
    if vdi_optimization.upper() in ("YES", "Y"):
        vdi_optimization = True
    else:
        vdi_optimization = False
    if clean_partitions.upper() in ("YES", "Y"):
        clean_partitions = True
    else:
        clean_partitions = False

    # Create API body
    cluster_storage_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'VdiOptimization':vdi_optimization,
        'DiskPartitionCleanup':clean_partitions
    }

    # Execute API
    try:
        api_response = cluster_storage_policy_handle.hyperflex_cluster_storage_policies_post(cluster_storage_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_auto_support_policy(api_instance, policy_name, auto_support, auto_support_email):

    # Get API handle
    auto_support_policy_handle = hyperflex_auto_support_policy_api.HyperflexAutoSupportPolicyApi(api_instance)

    # Setup variables
    if auto_support.upper() in ("YES", "Y"):
        auto_support = True
    else:
        auto_support = False

    # Create API body
    auto_support_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'ServiceTicketReceipient':auto_support_email,
        'AdminState':auto_support
    }

    # Execute API
    try:
        api_response = auto_support_policy_handle.hyperflex_auto_support_policies_post(auto_support_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_node_config_policy(api_instance, policy_name, hostname_prefix, mgmt_ip_range, hxdp_ip_range):

    # Get API handle
    node_config_policy_handle = hyperflex_node_config_policy_api.HyperflexNodeConfigPolicyApi(api_instance)

    # Setup variables

    # Create API body
    node_config_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'NodeNamePrefix':hostname_prefix,
        'MgmtIpRange':mgmt_ip_range,
        'HxdpIpRange':hxdp_ip_range
    }

    # Execute API
    try:
        api_response = node_config_policy_handle.hyperflex_node_config_policies_post(node_config_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_cluster_network_policy(api_instance, policy_name, mgmt_vlan, uplink_speed, jumbo_frames, mac_address_range):

    # Get API handle
    cluster_network_policy_handle = hyperflex_cluster_network_policy_api.HyperflexClusterNetworkPolicyApi(api_instance)

    # Setup variables
    if jumbo_frames.upper() in ("YES", "Y"):
        jumbo_frames = True
    else:
        jumbo_frames = False


    # Create API body
    cluster_network_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'MgmtVlan':mgmt_vlan,
        'UplinkSpeed':uplink_speed,
        'JumboFrame':jumbo_frames,
        'MacPrefixRange':mac_address_range
    }

    # Execute API
    try:
        api_response = cluster_network_policy_handle.hyperflex_cluster_network_policies_post(cluster_network_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_proxy_setting_policy(api_instance, policy_name, hostname, port, username, password):

    # Get API handle
    proxy_setting_policy_handle = hyperflex_proxy_setting_policy_api.HyperflexProxySettingPolicyApi(api_instance)

    # Setup variables

    # Create API body

    proxy_setting_policy_body = {
        'Name':policy_name,
        #'Description':description,
        'Hostname':hostname,
        'Port':port,
        'Username':username,
        'Password':password
    }

    # Execute API
    try:
        api_response = proxy_setting_policy_handle.hyperflex_proxy_setting_policies_post(proxy_setting_policy_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_hx_ip_range(end_ip, gw_ip, subnet_mask, start_ip):
    ip_range = hyperflex_ip_addr_range.HyperflexIpAddrRange(end_ip, gw_ip, subnet_mask, start_ip)
    return ip_range


def create_hx_vlan(vlan_name, vlan_id):
    vlan = hyperflex_named_vlan.HyperflexNamedVlan(vlan_name, int(vlan_id))
    return vlan

def create_hx_mac_addr_prefix_range(end_addr, start_addr):
    mac_address_range = hyperflex_mac_addr_prefix_range.HyperflexMacAddrPrefixRange(end_addr, start_addr)
    return mac_address_range

def hyperflex_cluster_profile_exists(api_instance, profile_name):

    # Get API handle
    hx_profiles_handle = hyperflex_cluster_profile_api.HyperflexClusterProfileApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % profile_name)

    # Execute API
    try:
        api_response = hx_profiles_handle.hyperflex_cluster_profiles_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def software_version_policy_exists(api_instance, policy_name):

    # Get API handle
    software_version_policy_handle = hyperflex_software_version_policy_api.HyperflexSoftwareVersionPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = software_version_policy_handle.hyperflex_software_version_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def credential_policy_exists(api_instance, policy_name):

    # Get API handle
    credential_policy_handle = hyperflex_local_credential_policy_api.HyperflexLocalCredentialPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = credential_policy_handle.hyperflex_local_credential_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def sys_config_policy_exists(api_instance, policy_name):

    # Get API handle
    sys_config_policy_handle = hyperflex_sys_config_policy_api.HyperflexSysConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = sys_config_policy_handle.hyperflex_sys_config_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)



def vcenter_policy_exists(api_instnace, policy_name):

    # Get API handle
    vcenter_config_policy_handle = hyperflex_vcenter_config_policy_api.HyperflexVcenterConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = vcenter_config_policy_handle.hyperflex_vcenter_config_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def cluster_storage_policy_exists(api_instance, policy_name):

    # Get API handle
    cluster_storage_policy_handle = hyperflex_cluster_storage_policy_api.HyperflexClusterStoragePolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = cluster_storage_policy_handle.hyperflex_cluster_storage_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def auto_support_policy_exists(api_instance, policy_name):

    # Get API handle
    auto_support_policy_handle = hyperflex_auto_support_policy_api.HyperflexAutoSupportPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = auto_support_policy_handle.hyperflex_auto_support_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def node_config_policy_exists(api_instance, policy_name):

    # Get API handle
    node_config_policy_handle = hyperflex_node_config_policy_api.HyperflexNodeConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = node_config_policy_handle.hyperflex_node_config_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def cluster_network_policy_exists(api_instance, policy_name):

    # Get API handle
    cluster_network_policy_handle = hyperflex_cluster_network_policy_api.HyperflexClusterNetworkPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = cluster_network_policy_handle.hyperflex_cluster_network_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def proxy_setting_policy_exists(api_instance, policy_name):

    # Get API handle
    proxy_setting_policy_handle = hyperflex_proxy_setting_policy_api.HyperflexProxySettingPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = proxy_setting_policy_handle.hyperflex_proxy_setting_policies_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def hx_node_profile_exists(api_instance, hx_node_moid):

    # Get API handle
    hx_node_profile_handle = hyperflex_node_profile_api.HyperflexNodeProfileApi(api_instance)

    # Setup variables
    kwargs = dict(filter="AssignedServer.Moid eq '%s'" % hx_node_moid)

    # Execute API
    try:
        api_response = hx_node_profile_handle.hyperflex_node_profiles_get(**kwargs)
        #print(api_response)
        if api_response.results == None:
            return False
        else:
            return True
    except ApiException as e:
        print("Error: %s\n" % e)


def get_software_version_policy_moid(api_instance, policy_name):

    # Get API handle
    software_version_policy_handle = hyperflex_software_version_policy_api.HyperflexSoftwareVersionPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = software_version_policy_handle.hyperflex_software_version_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_credential_policy_moid(api_instance, policy_name):

    # Get API handle
    credential_policy_handle = hyperflex_local_credential_policy_api.HyperflexLocalCredentialPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = credential_policy_handle.hyperflex_local_credential_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_sys_config_policy_moid(api_instance, policy_name):

    # Get API handle
    sys_config_policy_handle = hyperflex_sys_config_policy_api.HyperflexSysConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = sys_config_policy_handle.hyperflex_sys_config_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)



def get_vcenter_policy_moid(api_instnace, policy_name):

    # Get API handle
    vcenter_config_policy_handle = hyperflex_vcenter_config_policy_api.HyperflexVcenterConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = vcenter_config_policy_handle.hyperflex_vcenter_config_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_cluster_storage_policy_moid(api_instance, policy_name):

    # Get API handle
    cluster_storage_policy_handle = hyperflex_cluster_storage_policy_api.HyperflexClusterStoragePolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = cluster_storage_policy_handle.hyperflex_cluster_storage_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_auto_support_policy_moid(api_instance, policy_name):

    # Get API handle
    auto_support_policy_handle = hyperflex_auto_support_policy_api.HyperflexAutoSupportPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = auto_support_policy_handle.hyperflex_auto_support_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_node_config_policy_moid(api_instance, policy_name):

    # Get API handle
    node_config_policy_handle = hyperflex_node_config_policy_api.HyperflexNodeConfigPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = node_config_policy_handle.hyperflex_node_config_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_cluster_network_policy_moid(api_instance, policy_name):

    # Get API handle
    cluster_network_policy_handle = hyperflex_cluster_network_policy_api.HyperflexClusterNetworkPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = cluster_network_policy_handle.hyperflex_cluster_network_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_proxy_setting_policy_moid(api_instance, policy_name):

    # Get API handle
    proxy_setting_policy_handle = hyperflex_proxy_setting_policy_api.HyperflexProxySettingPolicyApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % policy_name)

    # Execute API
    try:
        api_response = proxy_setting_policy_handle.hyperflex_proxy_setting_policies_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_hyperflex_cluster_profile_moid(api_instance, profile_name):

    # Get API handle
    hx_profiles_handle = hyperflex_cluster_profile_api.HyperflexClusterProfileApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Name eq '%s'" % profile_name)

    # Execute API
    try:
        api_response = hx_profiles_handle.hyperflex_cluster_profiles_get(**kwargs)
        #print(api_response)
        return api_response.results[0].moid
    except ApiException as e:
        print("Error: %s\n" % e)


def get_compute_rack_unit_moid_by_serial(api_instance, serial_num):

    # Get API handle
    compute_rack_unit_handle = compute_rack_unit_api.ComputeRackUnitApi(api_instance)

    # Setup variables
    kwargs = dict(filter="Serial eq '%s'" % serial_num)

    # Execute API
    try:
        api_response = compute_rack_unit_handle.compute_rack_units_get(**kwargs)
        #print(api_response)
        if api_response.results != None:
            return api_response.results[0].moid
        else:
            return None
    except ApiException as e:
        print("Error: %s\n" % e)


def create_hx_cluster_profile(api_instance, profile_name, data_vlan, cluster_mgmt_ip, hxdp_version, mac_address_prefix, record_policy_map):

    # Get API handle
    hx_cluster_profile_handle = hyperflex_cluster_profile_api.HyperflexClusterProfileApi(api_instance)

    # Setup variables
    software_version_policy_moid = record_policy_map['software_version_policy']
    credential_policy_moid = record_policy_map['credential_policy']
    sys_config_policy_moid = record_policy_map['sys_config_policy']
    vcenter_policy_moid = record_policy_map['vcenter_policy']
    cluster_storage_policy_moid = record_policy_map['cluster_storage_policy']
    auto_support_policy_moid = record_policy_map['auto_support_policy']
    node_config_policy_moid = record_policy_map['node_config_policy']
    cluster_network_policy_moid = record_policy_map['cluster_network_policy']
    proxy_setting_policy_moid = record_policy_map['proxy_setting_policy']

    hx_cluster_profile_body = {
        'Name':profile_name,                            #General-->Name
        #'HxdpVersion':hxdp_version,                     #General-->HyperFlex Data Platform Version
        'SoftwareVersion':software_version_policy_moid,
        'MgmtPlatform':'EDGE',                          #General-->Type (i.e. Edge vs HX w/ FIs)
        'Replication':2,                                #General-->Replication Factor
        'Description':'',                               #General-->Description
        #'Tags':'',                                     #General-->Add Tags
        'LocalCredential':credential_policy_moid,       #Cluster Configuration-->Security policy
        'SysConfig':sys_config_policy_moid,             #Cluster Configuration-->DNS, NTP and Timezone policy
        'VcenterConfig':vcenter_policy_moid,            #Cluster Configuration-->vCenter policy
        'ClusterStorage':cluster_storage_policy_moid,   #Cluster Configuration-->Storage Configuration
        'AutoSupport':auto_support_policy_moid,         #Cluster Configuration-->Auto Support policy
        'NodeConfig':node_config_policy_moid,           #Cluster Configuration-->IP & Hostname policy
        #'UcsmConfig':'',                               #Cluster Configuration-->UCS Manager Configuration policy
        'ClusterNetwork':cluster_network_policy_moid,   #Cluster Configuration-->Network Configuration policy
        'ProxySetting':proxy_setting_policy_moid,       #Cluster Configuration-->Proxy Setting
        'StorageDataVlan':data_vlan,                    #Cluster Configuration-->HyperFlex Storage Network
        'MgmtIpAddress':cluster_mgmt_ip,                #Nodes Configuration-->Cluster Management IP Address
        'MacAddressPrefix':mac_address_prefix,          #Nodes Configuration-->MAC Prefix Address
    }

    # Execute API
    try:
        api_response = hx_cluster_profile_handle.hyperflex_cluster_profiles_post(hx_cluster_profile_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)



def claim_imc_device(api_instance, claim_code, device_id):

    # Get API handle
    claim_handle = asset_device_claim_api.AssetDeviceClaimApi(api_instance)

    # Setup variables
    claim_body = {
    'SecurityToken': claim_code,
    'SerialNumber': device_id
    }

    # Execute Api
    try:
        api_response = claim_handle.asset_device_claims_post(claim_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def create_hx_node_profile(api_instance, hx_node_moid, hx_cluster_profile_moid, hx_node_profile_name):

    # Get API handle
    hx_node_profile_handle = hyperflex_node_profile_api.HyperflexNodeProfileApi(api_instance)

    # Setup variables
    hx_node_profile_body = {
        'AssignedServer':hx_node_moid,
        'ClusterProfile':hx_cluster_profile_moid,
        'Name':hx_node_profile_name
    }

    # Execute API
    try:
        api_response = hx_node_profile_handle.hyperflex_node_profiles_post_with_http_info(hx_node_profile_body)
        #print(api_response)
    except ApiException as e:
        print("Error: %s\n" % e)


def timeout_handler(num, stack):
    raise Exception('timedout')


####################################################################################################
#
#
# LOGGING SETUP
#
#
####################################################################################################

current_time = datetime.datetime.now()
logfile = './logs/'+str(current_time)
format = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(filename=logfile,level=logging.DEBUG,format=format)
format = '%(asctime)s - %(levelname)s - %(message)s'



####################################################################################################
#
#
# SCRIPT ARGUMENT SETUP
#
#
####################################################################################################

parser = argparse.ArgumentParser()
help_str = 'JSON file with Intersight API parameters.  Default: intersight_api_params.json'
parser.add_argument('-a', '--api_params', default='./inputs/auth', help=help_str)
parser.add_argument('-f', '--input_file', default='./inputs/input.xlsx')
args = parser.parse_args()





####################################################################################################
#
#
# MENU PROMPT
#
#
####################################################################################################

print("\n")
print(Fore.CYAN+"Please select from the following options:\n"+Style.RESET_ALL)
while True:
    print("     1. Claim HyperFlex nodes in Intersight\n \
    2. Create HyperFlex Cluster Profiles in Intersight\n \
    3. Assign claimed HyperFlex nodes to HyperFlex Cluster Profiles in Intersight\n \
    4. Perform all of the above\n")

    menu_answer = raw_input("     Enter the number for your selection: ")
    if menu_answer in ('1','2','3','4'):
        break
    else:
        print(Fore.CYAN+"Not a valid option, please select again: \n"+Style.RESET_ALL)




####################################################################################################
#
#
# GET PASSWORDS VIA CLI PROMPT
#
#
####################################################################################################

print("\n")
print(Fore.CYAN+"Collecting required passwords...\n"+Style.RESET_ALL)

if menu_answer in ('2','4'):
    while True:
        hypervisor_password = getpass.getpass("     Please enter the new ESXi hypervisor password: ")
        confirm_hypervisor_password = getpass.getpass("     Please confirm the new ESXi hypervisor password: ")
        if hypervisor_password == confirm_hypervisor_password:
            break
        else:
            print (Fore.RED+"        Passwords do not match, please re-enter the passwords."+Style.RESET_ALL)

    pass_answer1 = raw_input("     Do you want the HyperFlex cluster to use the same password? (yes/no): ")
    if pass_answer1.upper() in ("YES", "Y"):
        controller_password = hypervisor_password
    else:
        while True:
            controller_password = getpass.getpass("        Please enter the new HyperFlex controller VM password: ")
            confirm_controller_password = getpass.getpass("        Please confirm the new HyperFlex controller VM password: ")
            if controller_password == confirm_controller_password:
                break
            else:
                print (Fore.RED+"      Passwords do not match, please re-enter the passwords."+Style.RESET_ALL)

    pass_answer2 = raw_input("     Is the vCenter password the same? (yes/no): ")
    if pass_answer2.upper() in ("YES", "Y"):
        vcenter_password = hypervisor_password
    else:
        vcenter_password = getpass.getpass("        Please enter the vCenter password: ")

if menu_answer in ('1','3','4'):
    cimc_password = getpass.getpass("     Please enter the CIMC password for the HyperFlex nodes: ")

print("\n")






####################################################################################################
#
#
# GET INTERSIGHT API INSTANCE
#
#
####################################################################################################

with open(args.api_params, 'r') as api_file:
    intersight_api_params = json.load(api_file)
api_instance = IntersightApiClient(
    host=intersight_api_params['api_base_uri'],
    private_key=intersight_api_params['api_private_key_file'],
    api_key_id=intersight_api_params['api_key_id'],
    )





####################################################################################################
#
#
# MAIN
#
#
####################################################################################################

# Setup output table header row
t = BeautifulTable(max_width=400)
if menu_answer == '1':
    t.column_headers = [Fore.CYAN+'                                                          HyperFlex Nodes                                                          '+Style.RESET_ALL]
elif menu_answer == '2':
    t.column_headers = [Fore.CYAN+'        HyperFlex Cluster Profiles        '+Style.RESET_ALL]
elif menu_answer in ('3','4'):
    t.column_headers = [Fore.CYAN+'        HyperFlex Cluster Profiles        '+Style.RESET_ALL, Fore.CYAN+'                                                          HyperFlex Nodes                                                          '+Style.RESET_ALL]
# Count required to properly format output table
count = 1

# Open input XLSX file
wb = load_workbook(args.input_file)
ws = wb.get_sheet_by_name('input')
#Iterate over each row in input CSV file
for row in ws.iter_rows(min_row=3, values_only=True):

    # Create new InputRecord
    new_record = InputRecord()
    new_record.hx_profile_name = row[0]
    new_record.cluster_type = row[1]
    new_record.hxdp_version = row[2]
    new_record.description = row[3]
    new_record.data_vlan_id = row[4]
    new_record.cluster_mgmt_ip = row[5]
    new_record.mac_address_prefix = row[6]
    new_record.hx_nodes_cimc_ips = row[7]
    new_record.hx_nodes_cimc_user = row[8]
    new_record.local_credential_policy_name = row[9]
    new_record.hypervisor_admin_user = row[10]
    new_record.sys_config_policy_name = row[11]
    new_record.timezone = row[12]
    new_record.dns_suffix = row[13]
    new_record.dns_servers = row[14]
    new_record.ntp_servers = row[15]
    new_record.vcenter_policy_name = row[16]
    new_record.vcenter = row[17]
    new_record.vcenter_user = row[18]
    new_record.vcenter_dc = row[19]
    new_record.vcenter_sso = row[20]
    new_record.cluster_storage_policy_name = row[21]
    new_record.vdi_optimization = row[22]
    new_record.clean_partitions = row[23]
    new_record.auto_support_policy_name = row[24]
    new_record.auto_support = row[25]
    new_record.auto_support_email = row[26]
    new_record.node_config_policy_name = row[27]
    new_record.hostname_prefix = row[28]
    new_record.mgmt_start_ip = row[29]
    new_record.mgmt_end_ip = row[30]
    new_record.mgmt_subnet_mask = row[31]
    new_record.mgmt_gw = row[32]
    new_record.cont_vm_start_ip = row[33]
    new_record.cont_vm_end_ip = row[34]
    new_record.cont_vm_subnet_mask = row[35]
    new_record.cont_vm_gw = row[36]
    new_record.cluster_network_policy_name = row[37]
    new_record.mgmt_vlan_id = row[38]
    new_record.uplink_speed = row[39]
    new_record.jumbo_frames = row[40]
    new_record.proxy_setting_policy_name = row[41]
    new_record.proxy_hostname = row[42]
    new_record.proxy_port = row[43]
    new_record.proxy_username = row[44]


    # Create hyperflex cluster profiles from input CSV if option 2 or option 3 is selected
    if menu_answer in ('2','4'):

        # Check if hyperflex cluster profile already exists in Intersight
        if hyperflex_cluster_profile_exists(api_instance, new_record.hx_profile_name):
            profile_status = 'exists'
        else:
            # Create dictionary to record policy moids for this hyperflex cluster profile instance
            record_policy_map = {}


            # SOFTWARE VERSION POLICY
            # Check if software version policy exists
            #if not software_version_policy_exists(api_instance, new_record.software_version_policy_name):
                # Create software version policy if does not exist
                #create_software_version_policy(api_instance, new_record.software_version_policy_name, new_record.hxdp_version)
            create_software_version_policy(api_instance, new_record.hx_profile_name, new_record.hxdp_version)
            record_policy_map['software_version_policy'] = get_software_version_policy_moid(api_instance, new_record.hx_profile_name)


            # CREDENTIAL POLICY
            # Check if local credential policy exists
            if not credential_policy_exists(api_instance, new_record.local_credential_policy_name):
                # Create local credential policy if does not exist
                create_credential_policy(api_instance, new_record.local_credential_policy_name, new_record.hypervisor_admin_user, hypervisor_password, controller_password)
            # Add credential policy moid to dictionary
            record_policy_map['credential_policy'] = get_credential_policy_moid(api_instance, new_record.local_credential_policy_name)


            # SYS CONFIG POLICY
            # Check if sys config policy exists
            if not sys_config_policy_exists(api_instance, new_record.sys_config_policy_name):
                # Create sys config policy if does not exist
                create_sys_config_policy(api_instance, new_record.sys_config_policy_name, new_record.timezone, new_record.dns_suffix, new_record.dns_servers, new_record.ntp_servers)
            # Add sys config policy moid to dictionary
            record_policy_map['sys_config_policy'] = get_sys_config_policy_moid(api_instance, new_record.sys_config_policy_name)


            # VCENTER POLICY
            # vCenter policy is optional, check if data in input CSV
            if new_record.vcenter_policy_name != None:
                # Check if vCenter policy exists
                if not vcenter_policy_exists(api_instance, new_record.vcenter_policy_name):
                    # Create vCenter policy if does not exist
                    create_vcenter_policy(api_instance, new_record.vcenter_policy_name, new_record.vcenter, new_record.vcenter_user, vcenter_password, new_record.vcenter_dc, new_record.vcenter_sso)
                # Add vCenter policy moid to dictionary
                record_policy_map['vcenter_policy'] = get_vcenter_policy_moid(api_instance, new_record.vcenter_policy_name)
            else:
                # No vCenter policy data in input CSV, skipping vCenter policy
                record_policy_map['vcenter_policy'] = None


            # CLUSTER STORAGE POLICY
            # Cluster storage policy is optional, check if data in input CSV
            if new_record.cluster_storage_policy_name != None:
                # Check if cluster storage policy exists
                if not cluster_storage_policy_exists(api_instance, new_record.cluster_storage_policy_name):
                    # Create cluster storage policy if does not exist
                    create_cluster_storage_policy(api_instance, new_record.cluster_storage_policy_name, new_record.vdi_optimization, new_record.clean_partitions)
                # Add cluster storage policy moid to dictionary
                record_policy_map['cluster_storage_policy'] = get_cluster_storage_policy_moid(api_instance, new_record.cluster_storage_policy_name)
            else:
                # No cluster storage policy data in input CSV, skipping cluster storage policy
                record_policy_map['cluster_storage_policy'] = None


            # AUTO SUPPORT POLICY
            # Auto support policy is optional, check if data in input CSV
            if new_record.auto_support_policy_name != None:
                # Check if auto support policy exists
                if not auto_support_policy_exists(api_instance, new_record.auto_support_policy_name):
                    # Create auto support policy if does not exist
                    create_auto_support_policy(api_instance, new_record.auto_support_policy_name, new_record.auto_support, new_record.auto_support_email)
                # Add auto support policy moid to dictionary
                record_policy_map['auto_support_policy'] = get_auto_support_policy_moid(api_instance, new_record.auto_support_policy_name)
            else:
                # No auto support policy data in input CSV, skipping auto support policy
                record_policy_map['auto_support_policy'] = None


            # NODE CONFIG POLICY
            # Check if node config policy exists
            if not node_config_policy_exists(api_instance, new_record.node_config_policy_name):
                # Create node config policy if does not exist
                mgmt_ip_range = create_hx_ip_range(new_record.mgmt_end_ip, new_record.mgmt_gw, new_record.mgmt_subnet_mask, new_record.mgmt_start_ip)
                cont_vm_ip_range = create_hx_ip_range(new_record.cont_vm_end_ip, new_record.cont_vm_gw, new_record.cont_vm_subnet_mask, new_record.cont_vm_start_ip)
                create_node_config_policy(api_instance, new_record.node_config_policy_name, new_record.hostname_prefix, mgmt_ip_range, cont_vm_ip_range)
            # Add node config policy moid to dictionary
            record_policy_map['node_config_policy'] = get_node_config_policy_moid(api_instance, new_record.node_config_policy_name)

            # NETWORK CONFIG POLICY
            # Check if cluster network policy exists
            if not cluster_network_policy_exists(api_instance, new_record.cluster_network_policy_name):
                # Create cluster network policy if does not exist
                mgmt_vlan = create_hx_vlan('mgmt', new_record.mgmt_vlan_id)
                mac_address_range = create_hx_mac_addr_prefix_range(new_record.mac_address_prefix, new_record.mac_address_prefix)
                create_cluster_network_policy(api_instance, new_record.cluster_network_policy_name, mgmt_vlan, new_record.uplink_speed ,new_record.jumbo_frames, mac_address_range)

            record_policy_map['cluster_network_policy'] = get_cluster_network_policy_moid(api_instance, new_record.cluster_network_policy_name)

            if new_record.proxy_setting_policy_name != None:
                # Check if proxy setting policy exists
                if not proxy_setting_policy_exists(api_instance, new_record.proxy_setting_policy_name):
                    # Create proxy setting policy if does not exist
                    create_proxy_setting_policy(api_instance, new_record.proxy_setting_policy_name, new_record.proxy_hostname, new_record.proxy_port, new_record.proxy_username, new_record.proxy_password)

                record_policy_map['proxy_setting_policy'] = get_proxy_setting_policy_moid(api_instance, new_record.proxy_setting_policy_name)
            else:
                record_policy_map['proxy_setting_policy'] = None


            # Create HyperFlex cluster profile
            data_vlan = create_hx_vlan('data', new_record.data_vlan_id)
            create_hx_cluster_profile(api_instance, new_record.hx_profile_name, data_vlan, new_record.cluster_mgmt_ip, new_record.hxdp_version, new_record.mac_address_prefix, record_policy_map)
            profile_status = 'created'



    # Claim devices if option 1 or option 3 is selected
    if menu_answer in ('1','4'):

        # Check if CIMC IP data exists for row in input CSV
        if new_record.hx_nodes_cimc_ips is not None:

            # Setup claim_status dictionary to record claim status for each hyperflex node specfified in input CSV row
            claim_status = {}

            # split hyperflex nodes into array
            hx_nodes_cimc_ip_array = []
            hx_nodes_cimc_ip_array = new_record.hx_nodes_cimc_ips.split(';')

            # make sure supported number of hyperflex nodes are provided for the row in input CSV
            if len(hx_nodes_cimc_ip_array) >= 2:

                # Iterate through array of hyperflex nodes
                for hx_node in hx_nodes_cimc_ip_array:

                    # Setup hyperflex node details in device dictionary
                    device = {}
                    device['device_type'] = 'imc'
                    device['hostname'] = hx_node
                    device['username'] = new_record.hx_nodes_cimc_user
                    device['password'] = cimc_password
                    device['read_only'] = False

                    # Proxy information is optional, check if proxy data exists in input CSV
                    if new_record.proxy_hostname is not None and new_record.proxy_port is not None:
                        device['proxy_host'] = new_record.proxy_hostname
                        device['proxy_port'] = new_record.proxy_port

                    # Setup timeout for 'try' statement below
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(20)

                    try:
                        # Get device_connector object for hyperflex node device
                        dc_obj = device_connector.UcsDeviceConnector(device)
                        if not dc_obj.logged_in:
                                dc_obj = device_connector.ImcDeviceConnector(device)

                        # Enable the device_connector for hyperflex node device if not already enabled, get response json
                        ro_json = dc_obj.configure_connector()

                    # Handle exception for incorrect CIMC password provided
                    except KeyError:
                        claim_status[hx_node] = 'not claimed'
                        logging.error('CLAIMING Device: '+hx_node+' --> Authentication to CIMC failed on node. Unable to get authentication cookie, check password and try again.')

                    # Handle exception for connection issues, i.e. timeout, etc.
                    except:
                        claim_status[hx_node] = 'not claimed'
                        logging.error('CLAIMING Device: '+hx_node+' --> Not able to connect to CIMC interface on node. Check IP address and connectivity and retry.')

                    # If 'try' statement above succeeds do the following
                    else:

                        # Set device_connector 'ReadOnlyMode' to False
                        if (ro_json.get('ReadOnlyMode') is not None) and (ro_json['ReadOnlyMode'] != device['read_only']):
                            ro_json = dc_obj.configure_access_mode(ro_json)

                        # Set device_connector proxy settings if required
                        if 'proxy_host' in device and 'proxy_port' in device:
                            ro_json = dc_obj.configure_proxy(ro_json, result)

                        # Wait for a connection to establish before checking claim state
                        for _ in range(10):
                            ro_json = dc_obj.get_status()
                            if ro_json['ConnectionState'] != 'Connected':
                                if ro_json['ConnectionState'] in ('DNS Misconfigured', 'Intersight DNS Resolve Error'):
                                    logging.error('CLAIMING Device: '+hx_node+' --> Check CIMC DNS settings, device connector reporting DNS misconfiguration and cannot connect to Intersight.')
                                    break
                                sleep(1)
                                ro_json = dc_obj.get_status()
                                #print ro_json
                            else:
                                break

                        if ro_json['ConnectionState'] != 'Connected':
                            claim_status[hx_node] = 'not claimed'
                            continue

                        if ro_json['AccountOwnershipState'] != 'Claimed':
                            # attempt to claim
                            (claim_resp, device_id, claim_code) = dc_obj.get_claim_info(ro_json)
                            claim_imc_device(api_instance, claim_code, device_id)
                            claim_status[hx_node] = 'claimed'
                        elif ro_json['AccountOwnershipState'] == 'Claimed':
                            claim_status[hx_node] = 'already claimed'
                        else:
                            claim_status[hx_node] = 'not claimed'
                        dc_obj.logout()

                    finally:
                        signal.alarm(0)



        # No CIMC data exists in row in input CSV
        else:
            # Set empty claim_status dictionary
            claim_status = {}


    # Assign nodes (devices) to HyperFlex cluster profile if option 3 or 4 selected
    if menu_answer in ('3','4'):
        if menu_answer == '4':
            sleep(30)

        # Setup claim_status dictionary to record claim status for each hyperflex node specfified in input CSV row
        assign_status = {}

        # Check if CIMC IP data exists for row in input CSV
        if new_record.hx_nodes_cimc_ips is not None:

            # split hyperflex nodes into array
            hx_nodes_cimc_ip_array = []
            hx_nodes_cimc_ip_array = new_record.hx_nodes_cimc_ips.split(';')

            # make sure supported number of hyperflex nodes are provided for the row in input CSV
            if len(hx_nodes_cimc_ip_array) >= 2:

                # Iterate through array of hyperflex nodes
                hx_node_serial_nums = []
                hx_node_count = 1
                for hx_node in hx_nodes_cimc_ip_array:

                    # Setup hyperflex node details in device dictionary
                    device = {}
                    device['hostname'] = hx_node
                    device['username'] = new_record.hx_nodes_cimc_user
                    device['password'] = cimc_password

                    # Setup timeout for 'try' statement below
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(20)

                    try:
                        # Get HX node serial number
                        imc_handle = ImcHandle(device['hostname'], device['username'], device['password'])
                        imc_handle.login()
                    except:
                        logging.error('ASSIGNING Device: '+hx_node+' --> Not able to log into CIMC interface.')
                    else:
                        object = imc_handle.query_dn("sys/rack-unit-1")
                        device['serial_num'] = object.serial
                        hx_node_serial_nums.append(device['serial_num'])
                        imc_handle.logout()
                    finally:
                        signal.alarm(0)

                    # Setup timeout for 'try' statement below
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(20)

                    try:
                        hx_node_moid = get_compute_rack_unit_moid_by_serial(api_instance, device['serial_num'])
                        if hx_node_moid == None:
                            raise Exception('Not claimed by Intersight. Device must first be claimed by Intersight before it can be assigned to a HyperFlex cluster profile.')
                        if hx_node_profile_exists(api_instance, hx_node_moid):
                            assign_status[hx_node] = 'already assigned'
                            continue
                        hx_cluster_profile_moid = get_hyperflex_cluster_profile_moid(api_instance, new_record.hx_profile_name)
                        if hx_cluster_profile_moid == None:
                            raise Exception('HyperFlex cluster profile does not exist. Check to make sure the HyperFlex cluster profile exists before trying to assign HyperFlex nodes.')
                        # Generate HyperFlex node profile name
                        hx_node_profile_name = new_record.hx_profile_name+"-"+str(hx_node_count)
                        create_hx_node_profile(api_instance, hx_node_moid, hx_cluster_profile_moid, hx_node_profile_name)

                    except Exception as e:
                        assign_status[hx_node] = 'unassigned'
                        error_msg = e.args
                        logging.error('ASSIGNING Device: '+hx_node+' --> '+str(error_msg))
                        continue
                    else:
                        assign_status[hx_node] = 'assigned'
                    finally:
                        signal.alarm(0)

                    hx_node_count += 1

                if('imc_handle' in locals() or 'imc_handle' in globals()):
                        imc_handle.logout()



    # Setup output table row if option 1 selected
    if menu_answer == '1':
        claim_status_display = ''
        if any(claim_status):
            for key, value in claim_status.iteritems():
                if value == 'claimed':
                    claim_status_display += key+"="+Fore.GREEN+value+Style.RESET_ALL+", "
                elif value == 'already claimed':
                    claim_status_display += key+"="+Fore.YELLOW+value+Style.RESET_ALL+", "
                elif value == 'not claimed':
                    claim_status_display += key+"="+Fore.RED+value+Style.RESET_ALL+", "
            claim_status_display = claim_status_display[:-2]
        t.append_row([claim_status_display])


    # Setup output table row if option 2 selected
    if menu_answer == '2':
        if profile_status == 'exists':
            profile_status_display = new_record.hx_profile_name+"="+Fore.YELLOW+'already exists'+Style.RESET_ALL
        elif profile_status == 'created':
            profile_status_display = new_record.hx_profile_name+"="+Fore.GREEN+'created'+Style.RESET_ALL
        else:
            profile_status_display = new_record.hx_profile_name+"="+Fore.RED+'error'+Style.RESET_ALL
        t.append_row([profile_status_display])


    if menu_answer == '3':
        profile_status_display = new_record.hx_profile_name
        assign_status_display = ''
        if any(assign_status):
            for key, value in assign_status.iteritems():
                if value == 'assigned':
                    assign_status_display += key+"="+Fore.GREEN+value+Style.RESET_ALL+", "
                elif value == 'already assigned':
                    assign_status_display += key+"="+Fore.YELLOW+value+Style.RESET_ALL+", "
                elif value == 'unassigned':
                    assign_status_display += key+"="+Fore.RED+value+Style.RESET_ALL+", "
            assign_status_display = assign_status_display[:-2]
        t.append_row([profile_status_display, assign_status_display])


    if menu_answer == '4':
        if profile_status == 'exists':
            profile_status_display = new_record.hx_profile_name+"="+Fore.YELLOW+'already exists'+Style.RESET_ALL
        elif profile_status == 'created':
            profile_status_display = new_record.hx_profile_name+"="+Fore.GREEN+'created'+Style.RESET_ALL
        else:
            profile_status_display = new_record.hx_profile_name+"="+Fore.RED+'not created'+Style.RESET_ALL
        node_status_display = ''
        if any(claim_status):
            for key, value in claim_status.iteritems():
                if value == 'claimed':
                    node_status_display += key+"="+Fore.GREEN+value+Style.RESET_ALL+":"
                elif value == 'already claimed':
                    node_status_display += key+"="+Fore.YELLOW+value+Style.RESET_ALL+":"
                elif value == 'not claimed':
                    node_status_display += key+"="+Fore.RED+value+Style.RESET_ALL+":"

                if assign_status[key] == 'assigned':
                    assign_decorator = Fore.GREEN
                elif assign_status[key] == 'already assigned':
                    assign_decorator = Fore.YELLOW
                else:
                    assign_decorator = Fore.RED
                node_status_display += assign_decorator+assign_status[key]+Style.RESET_ALL+", "

            node_status_display = node_status_display[:-2]
        t.append_row([profile_status_display, node_status_display])



    # Following lines required to make output table expand in real-time and display correctly
    if count == 1:
        print t
    else:
        print( "\n".join(t.get_string().splitlines()[-2:]))
    count += 1

print("\n")
